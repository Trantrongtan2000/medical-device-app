import sqlite3
import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

qltb_db = Path(r"G:\QLTB\database.db")
excel_path = Path(r"C:\Users\tantt\Downloads\CĐHA - Danh sách máy siêu âm 12-08-2026 NEW Q7.xlsx")

print("🔍 PHÂN TÍCH CHI TIẾT G:\\QLTB\\database.db VÀ TOÀN BỘ MÁY CĐHA Q7:\n" + "=" * 75)

# 1. Khám phá SQLite database.db trong G:\QLTB
if qltb_db.exists():
    print(f"📂 1. CSDL 'G:\\QLTB\\database.db' (Dung lượng: {qltb_db.stat().st_size/1024:.1f} KB):")
    conn = sqlite3.connect(qltb_db)
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [r[0] for r in cur.fetchall()]
    print(f"   Danh sách bảng ({len(tables)}): {tables}")
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        cnt = cur.fetchone()[0]
        cur.execute(f"PRAGMA table_info({t})")
        cols = [c[1] for c in cur.fetchall()]
        print(f"   • Bảng [{t}]: {cnt:,} dòng | Cột: {cols[:8]}")
    conn.close()

# 2. Phân tích chi tiết toàn bộ máy siêu âm & đầu dò trong Excel CĐHA Q7
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Sheet SÂ (Máy siêu âm và danh sách đầu dò)
ws_sa = wb['SÂ']
print(f"\n📊 2. SHEET [SÂ] - TỔNG HỢP MÁY SIÊU ÂM & ĐẦU DÒ CHI TIẾT TẠI CĐHA Q7:")
sa_machines = []
current_m = None

for r in range(3, ws_sa.max_row + 1):
    stt = ws_sa.cell(r, 1).value
    room = ws_sa.cell(r, 2).value
    m_name = ws_sa.cell(r, 3).value
    qty = ws_sa.cell(r, 4).value
    sn = ws_sa.cell(r, 5).value
    probe_name = ws_sa.cell(r, 6).value
    probe_qty = ws_sa.cell(r, 7).value
    probe_sn = ws_sa.cell(r, 8).value
    year = ws_sa.cell(r, 9).value
    notes = ws_sa.cell(r, 10).value

    if stt and (room or m_name):
        current_m = {
            "stt": stt,
            "room": str(room or '').strip(),
            "name": str(m_name or '').strip(),
            "serial_no": str(sn or '').strip().replace('.0', ''),
            "year": str(year or '').strip().replace('.0', ''),
            "notes": str(notes or '').strip(),
            "probes": [],
            "ups": None
        }
        sa_machines.append(current_m)

    if current_m:
        if m_name and "UPS" in str(m_name).upper():
            current_m["ups"] = str(sn or '').strip().replace('.0', '')
        if probe_name:
            current_m["probes"].append({
                "probe_name": str(probe_name).strip(),
                "probe_sn": str(probe_sn or '').strip().replace('.0', '')
            })

print(f"   ✅ Tổng số Máy Siêu Âm CĐHA Q7: {len(sa_machines)} hệ thống máy chính")
for idx, m in enumerate(sa_machines[:12], 1):
    print(f"   {idx:02d}. [Phòng {m['room']:12s}] {m['name']:25s} | SN: {m['serial_no']:15s} | {len(m['probes'])} Đầu dò (UPS: {m['ups'] or 'K/c'})")
    for p in m['probes']:
        print(f"       └── 🩺 {p['probe_name']} (SN: {p['probe_sn']})")

# Sheet XQCTMRINAĐLX (Chụp CHT 3T/1.5T, CT đa lát cắt, X-Quang, Đo loãng xương)
ws_xq = wb['XQCTMRINAĐLX']
print(f"\n📊 3. SHEET [XQCTMRINAĐLX] - TỔNG HỢP HỆ THỐNG MRI, CT, X-QUANG, ĐO LOÃNG XƯƠNG CĐHA Q7:")
xq_machines = []
for r in range(3, ws_xq.max_row + 1):
    stt = ws_xq.cell(r, 1).value
    room = ws_xq.cell(r, 2).value
    m_name = ws_xq.cell(r, 3).value
    qty = ws_xq.cell(r, 4).value
    sn = ws_xq.cell(r, 5).value
    year = ws_xq.cell(r, 6).value
    notes = ws_xq.cell(r, 7).value
    
    if stt and (room or m_name):
        xq_machines.append({
            "stt": stt,
            "room": str(room or '').strip().replace('.0', ''),
            "name": str(m_name or '').strip(),
            "serial_no": str(sn or '').strip().replace('.0', ''),
            "year": str(year or '')[:10],
            "notes": str(notes or '').strip()
        })

print(f"   ✅ Tổng số Hệ Thống Kỹ Thuật Cao (MRI, CT, XQ, ĐLX): {len(xq_machines)} hệ thống")
for idx, m in enumerate(xq_machines, 1):
    print(f"   {idx:02d}. [Phòng {m['room']:8s}] {m['name']:30s} | SN: {m['serial_no']:18s} | Ghi chú: {m['notes'] or 'Hoạt động'}")
