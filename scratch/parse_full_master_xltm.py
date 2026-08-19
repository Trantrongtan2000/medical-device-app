import openpyxl
import sys
from pathlib import Path
from collections import defaultdict, Counter

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712\Master Data.xltm")

print(f"🏥 PHÂN TÍCH TOÀN BỘ NỘI DUNG TỆP 'Master Data.xltm':\n" + "=" * 75)

wb = openpyxl.load_workbook(file_path, data_only=True)

# Sheet 1: Hop dong mua sam
ws1 = wb['1. Hop dong mua sam']
h1 = [ws1.cell(1, c).value for c in range(1, ws1.max_column + 1)]
valid_contracts = []
for r in range(2, ws1.max_row + 1):
    row_vals = [ws1.cell(r, c).value for c in range(1, len(h1) + 1)]
    if any(row_vals) and (row_vals[1] or row_vals[4]):
        valid_contracts.append(dict(zip([str(col or f'col_{idx}') for idx, col in enumerate(h1)], row_vals)))

print(f"📊 Sheet 1 [1. Hop dong mua sam]: {len(valid_contracts):,} bản ghi hợp đồng mua sắm hợp lệ")

# Sheet 2: Ban giao lap dat
ws2 = wb['2. Ban giao lap dat']
h2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
valid_handovers = []
dept_counts = Counter()
for r in range(2, ws2.max_row + 1):
    row_vals = [ws2.cell(r, c).value for c in range(1, len(h2) + 1)]
    if any(row_vals) and (row_vals[3] or row_vals[7]):
        row_dict = dict(zip([str(col or f'col_{idx}') for idx, col in enumerate(h2)], row_vals))
        valid_handovers.append(row_dict)
        dept = str(row_dict.get('Khoa') or 'Chưa rõ').strip()
        dept_counts[dept] += 1

print(f"📊 Sheet 2 [2. Ban giao lap dat]: {len(valid_handovers):,} thiết bị bàn giao lắp đặt chi tiết")
print(f"   Phân bổ theo Khoa phòng trong Sheet 2:")
for d, cnt in dept_counts.most_common(12):
    print(f"     • {d}: {cnt} thiết bị")

# Sheet 3: Bao tri
ws3 = wb['3. Bao tri']
h3 = [ws3.cell(1, c).value for c in range(1, ws3.max_column + 1)]
valid_maintenance = []
for r in range(2, ws3.max_row + 1):
    row_vals = [ws3.cell(r, c).value for c in range(1, len(h3) + 1)]
    if any(row_vals) and (row_vals[1] or row_vals[2]):
        valid_maintenance.append(dict(zip([str(col or f'col_{idx}') for idx, col in enumerate(h3)], row_vals)))

print(f"\n📊 Sheet 3 [3. Bao tri]: {len(valid_maintenance):,} kế hoạch bảo trì thiết bị")

# Sheet 4: Dropdown
ws4 = wb['Dropdown']
h4 = [ws4.cell(1, c).value for c in range(1, ws4.max_column + 1)]
print(f"\n📊 Sheet 4 [Dropdown]: Các danh mục Dropdown chuẩn hóa:")
for c_idx, col_name in enumerate(h4, 1):
    vals = [ws4.cell(r, c_idx).value for r in range(2, ws4.max_row + 1) if ws4.cell(r, c_idx).value]
    print(f"   • Danh mục [{col_name}]: {len(vals)} mục (Ví dụ: {', '.join([str(v) for v in vals[:3]])}...)")
