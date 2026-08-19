import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712\Master Data.xltm")

print(f"🔍 BẮT ĐẦU ĐỌC VÀ PHÂN TÍCH TỆP: {file_path}\n" + "=" * 70)

if not file_path.exists():
    # Try other locations where Master Data.xltm might exist
    alt_paths = list(Path(r"G:\BV QUẬN 7_OCR_WORK_20260712").rglob("*Master Data.xltm*"))
    print("❌ Không tìm thấy ở thư mục gốc! Các tệp tìm thấy:")
    for p in alt_paths:
        print(f"  • {p}")
    if alt_paths:
        file_path = alt_paths[0]
    else:
        sys.exit(1)

print(f"📂 Đang mở file: {file_path} (Dung lượng: {file_path.stat().st_size/1024:.1f} KB)")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
    print(f"📊 Danh sách các Sheet trong Workbook ({len(wb.sheetnames)} sheets):")
    for s_idx, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  {s_idx:02d}. Sheet: [{name}] (Kích thước: {ws.max_row} dòng x {ws.max_column} cột)")
    
    print("\n" + "=" * 70)
    
    # Read each sheet
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n📑 CHI TIẾT SHEET: [{name}] (Tối đa {ws.max_row} dòng):")
        
        # Read header row (find first non-empty row)
        headers = []
        header_row_idx = 1
        for r_idx in range(1, min(10, ws.max_row + 1)):
            row_vals = [ws.cell(r_idx, c).value for c in range(1, min(25, ws.max_column + 1))]
            if any(row_vals):
                headers = [str(v) if v is not None else f"Col_{i}" for i, v in enumerate(row_vals, 1)]
                header_row_idx = r_idx
                break
                
        print(f"   Dòng tiêu đề (Hàng {header_row_idx}): {headers[:12]}")
        print("   --- 5 dòng dữ liệu mẫu ---")
        for r_idx in range(header_row_idx + 1, min(header_row_idx + 6, ws.max_row + 1)):
            row_vals = [ws.cell(r_idx, c).value for c in range(1, len(headers) + 1)]
            if any(row_vals):
                row_str = " | ".join([str(v) if v is not None else "" for v in row_vals[:8]])
                print(f"   Row {r_idx:02d}: {row_str}")

except Exception as e:
    print(f"⚠️ Lỗi đọc file Excel/XLTM: {e}")
