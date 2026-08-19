import sqlite3
import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

print("🔍 TÌM KIẾM CHÍNH XÁC NHÀ CUNG CẤP AN VIỆT VÀ MÁY HERA W10:\n" + "=" * 75)

# 1. Tìm trong Master Data.xltm
xltm_path = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712\Master Data.xltm")
if xltm_path.exists():
    wb = openpyxl.load_workbook(xltm_path, data_only=True)
    ws1 = wb['1. Hop dong mua sam']
    print("📑 Tìm trong '1. Hop dong mua sam' của Master Data.xltm:")
    for r in range(2, ws1.max_row + 1):
        c_no = ws1.cell(r, 2).value
        item_name = ws1.cell(r, 3).value
        sup = ws1.cell(r, 4).value
        c_str = str(c_no or '')
        item_str = str(item_name or '')
        sup_str = str(sup or '')
        if "AN VIỆT" in sup_str.upper() or "AN VIET" in sup_str.upper() or "HERA" in item_str.upper() or "W10" in item_str.upper() or "VOLUSON" in item_str.upper():
            print(f"  • Row {r:03d} | HĐ: {c_str:30s} | Tên TB: {item_str:40s} | Nhà thầu: {sup_str}")

# 2. Tìm trong database devices.db
db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")
if db_path.exists():
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT id, device_name, model, serial_no, manufacturer FROM devices WHERE model LIKE '%Hera%' OR model LIKE '%W10%' OR device_name LIKE '%Hera%'")
    rows = cur.fetchall()
    print(f"\n📂 Tìm trong database devices.db ({len(rows)} kết quả):")
    for r in rows:
        print(f"  • ID: {r[0]} | Tên: {r[1]} | Model: {r[2]} | S/N: {r[3]} | Hãng: {r[4]}")
    conn.close()
