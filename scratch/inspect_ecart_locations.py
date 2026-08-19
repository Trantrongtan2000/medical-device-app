import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"C:\Users\tantt\Downloads\TA5. VỊ TRÍ KHOA PHÒNG - XE ECART.xlsx")

print(f"🔍 BẮT ĐẦU PHÂN TÍCH TỆP: {file_path.name}\n" + "=" * 75)

if not file_path.exists():
    print("❌ Không tìm thấy tệp!")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"📊 Danh sách Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

for s_name in wb.sheetnames:
    ws = wb[s_name]
    print(f"\n📑 SHEET: [{s_name}] (Kích thước: {ws.max_row} dòng x {ws.max_column} cột)")
    
    # In 20 dòng đầu tiên
    for r in range(1, min(25, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
        if any(row_vals):
            row_str = " | ".join([str(v) if v is not None else "" for v in row_vals[:12]])
            print(f"   Row {r:02d}: {row_str}")
