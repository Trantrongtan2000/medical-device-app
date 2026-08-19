import os
import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

qltb_path = Path(r"G:\QLTB")
excel_path = Path(r"C:\Users\tantt\Downloads\CĐHA - Danh sách máy siêu âm 12-08-2026 NEW Q7.xlsx")

print("🔍 BẮT ĐẦU KHÁM PHÁ DỮ LIỆU 'G:\\QLTB' VÀ TỆP DANH SÁCH MÁY SIÊU ÂM CĐHA:\n" + "=" * 75)

# 1. Khám phá G:\QLTB
print(f"📁 1. THƯ MỤC 'G:\\QLTB':")
if qltb_path.exists():
    items = list(qltb_path.iterdir())
    print(f"   Tìm thấy {len(items)} mục trong G:\\QLTB:")
    for idx, item in enumerate(items, 1):
        if item.is_dir():
            sub_count = len(list(item.iterdir())) if item.is_dir() else 0
            print(f"   {idx:02d}. [DIR]  {item.name} ({sub_count} mục con)")
        else:
            print(f"   {idx:02d}. [FILE] {item.name} ({item.stat().st_size/1024:.1f} KB)")
else:
    print("   ❌ Không tìm thấy đường dẫn G:\\QLTB")

# 2. Khám phá tệp Excel Máy siêu âm CĐHA
print(f"\n📊 2. TỆP EXCEL: {excel_path.name}")
if excel_path.exists():
    print(f"   Dung lượng: {excel_path.stat().st_size/1024:.1f} KB")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"   Danh sách Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    
    for s_name in wb.sheetnames:
        ws = wb[s_name]
        print(f"\n   📑 SHEET: [{s_name}] (Kích thước: {ws.max_row} dòng x {ws.max_column} cột)")
        
        # In các dòng đầu
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
            if any(row_vals):
                row_str = " | ".join([str(v) if v is not None else "" for v in row_vals[:10]])
                print(f"     Row {r:02d}: {row_str}")
else:
    print(f"   ❌ Không tìm thấy tệp: {excel_path}")
