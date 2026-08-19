import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"C:\Users\tantt\Downloads\TA5. VỊ TRÍ KHOA PHÒNG - XE ECART.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"📊 DANH SÁCH TOÀN BỘ 14 SHEETS TRONG '{file_path.name}':\n" + "=" * 75)

for idx, s_name in enumerate(wb.sheetnames, 1):
    ws = wb[s_name]
    # Get first row title
    title = ws.cell(1, 1).value or ws.cell(2, 1).value or ws.cell(1, 2).value or "N/A"
    print(f"  {idx:02d}. [{s_name:22s}] ({ws.max_row:4d} dòng x {ws.max_column:2d} cột) — Tiêu đề: {str(title)[:50]}")
