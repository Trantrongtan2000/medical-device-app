import sqlite3
import openpyxl
import csv
import re
import sys
from pathlib import Path
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")
ocr_root = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712")

print("🔍 BẮT ĐẦU ĐỐI CHỨNG DỮ LIỆU ĐỂ RÀ SOÁT SAI LỆCH TOÀN DIỆN:\n" + "=" * 70)

conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# 1. Đọc toàn bộ thiết bị trong SQLite Database
cur.execute("""
    SELECT d.id, d.device_name, d.model, d.serial_no, d.facility_id, d.risk_level, d.status,
           d.calibration_date, d.recalibration_date, d.certification_no, d.calibration_stamp_no,
           f.name as facility_name
    FROM devices d
    LEFT JOIN facilities f ON d.facility_id = f.id
""")
db_devices = [dict(r) for r in cur.fetchall()]
db_by_serial = {d['serial_no'].strip().upper(): d for d in db_devices if d.get('serial_no')}
db_by_id = {d['id']: d for d in db_devices}

print(f"📊 Dữ liệu hiện tại trong SQLite DB: {len(db_devices):,} thiết bị")

# 2. Đọc tệp Master kiểm định Markdown: Master_kiem_dinh_TB.md
md_master_path = ocr_root / "md" / "05_KIEM DINH" / "pdf" / "Master_kiem_dinh_TB.md"
md_records = []
if md_master_path.exists():
    with open(md_master_path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip().startswith('|') and '---' not in line and 'Khoa /Phòng' not in line:
                cols = [c.strip() for c in line.strip().split('|')[1:-1]]
                if len(cols) >= 12:
                    md_records.append({
                        "facility": cols[0],
                        "name": cols[1],
                        "model": cols[2],
                        "serial_no": cols[3],
                        "reg_no": cols[4],
                        "cert_type": cols[5],
                        "vendor": cols[6],
                        "cert_no": cols[7],
                        "stamp_no": cols[8],
                        "from_date": cols[9],
                        "to_date": cols[10],
                        "status": cols[11] if len(cols) > 11 else ""
                    })

print(f"📊 Dữ liệu từ Master_kiem_dinh_TB.md: {len(md_records):,} bản ghi kiểm định")

# 3. Đọc tệp Master Bàn Giao CSV: handover_master_enriched.csv & device_registry.csv
handover_csv_path = ocr_root / "03_BAN_GIAO_VA_NGHIEM_THU" / "_ocr_handover_assets" / "handover_master_enriched.csv"
handover_records = []
if handover_csv_path.exists():
    with open(handover_csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        handover_records = list(reader)

print(f"📊 Dữ liệu từ handover_master_enriched.csv: {len(handover_records):,} bản ghi bàn giao")

# 4. Đọc tệp Excel Master: 30.10.2024 Master Q7.xlsx nếu có
excel_master_path = ocr_root / "04_KIEM_DINH_VA_HIEU_CHUAN" / "2024" / "CÁ NHÂN" / "Tài" / "30.10.2024 Master Q7.xlsx"
excel_records = []
if excel_master_path.exists():
    try:
        wb = openpyxl.load_workbook(excel_master_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                excel_records.append(dict(zip([str(h) for h in headers], row)))
        print(f"📊 Dữ liệu từ Excel '30.10.2024 Master Q7.xlsx': {len(excel_records):,} dòng")
    except Exception as e:
        print(f"⚠️ Không đọc được Excel: {e}")

# ==================== PHÂN TÍCH ĐỐI CHỨNG SAI LỆCH ====================
print("\n" + "=" * 70)
print("⚖️ KẾT QUẢ ĐỐI CHỨNG DỮ LIỆU & RÀ SOÁT SAI LỆCH:")
print("=" * 70)

discrepancies = {
    "missing_serial_in_db": [],
    "date_mismatches": [],
    "cert_no_mismatches": [],
    "facility_mismatches": [],
    "status_mismatches": [],
    "unassigned_facility_in_db": []
}

# Check unassigned facilities in DB
for d in db_devices:
    if not d.get('facility_id') or d.get('facility_name') == 'Khoa/Phòng Chưa Phân Loại' or not d.get('facility_name'):
        discrepancies["unassigned_facility_in_db"].append(d)

# Cross-check Markdown Master vs DB
for r in md_records:
    sn = r['serial_no'].strip().upper()
    if not sn or sn == 'N/A' or sn == '-':
        continue
        
    if sn not in db_by_serial:
        discrepancies["missing_serial_in_db"].append(r)
    else:
        db_item = db_by_serial[sn]
        
        # Check Cert No
        if r['cert_no'] and db_item.get('certification_no'):
            if r['cert_no'].strip().lower() not in db_item['certification_no'].strip().lower():
                discrepancies["cert_no_mismatches"].append({
                    "serial": sn,
                    "db_cert": db_item['certification_no'],
                    "master_cert": r['cert_no'],
                    "name": db_item['device_name']
                })
                
        # Check Recalibration Date
        if r['to_date'] and db_item.get('recalibration_date'):
            master_date = r['to_date'].strip()
            db_date = str(db_item['recalibration_date']).strip()
            if master_date != db_date:
                discrepancies["date_mismatches"].append({
                    "serial": sn,
                    "name": db_item['device_name'],
                    "db_date": db_date,
                    "master_date": master_date,
                    "facility": db_item['facility_name']
                })
                
        # Check Facility name
        if r['facility'] and db_item.get('facility_name'):
            f_master = r['facility'].strip().upper()
            f_db = db_item['facility_name'].strip().upper()
            if f_master not in f_db and f_db not in f_master and f_db != "KHOA/PHÒNG CHƯA PHÂN LOẠI":
                discrepancies["facility_mismatches"].append({
                    "serial": sn,
                    "name": db_item['device_name'],
                    "db_fac": db_item['facility_name'],
                    "master_fac": r['facility']
                })

print(f"\n1. 🔍 Sai lệch thiết bị có trong Master Kiểm Định nhưng chưa có trong DB:")
print(f"   • Số lượng: {len(discrepancies['missing_serial_in_db'])} máy")
for m in discrepancies['missing_serial_in_db'][:5]:
    print(f"     - [{m['name']}] Model: {m['model']} | SN: {m['serial_no']} | Khoa: {m['facility']} | Hạn KĐ: {m['to_date']}")

print(f"\n2. 🔍 Sai lệch Hạn Kiểm Định (Recalibration Date Discrepancies):")
print(f"   • Số lượng: {len(discrepancies['date_mismatches'])} máy")
for m in discrepancies['date_mismatches'][:5]:
    print(f"     - [{m['name']}] SN: {m['serial']} | DB: {m['db_date']} vs Master: {m['master_date']} ({m['facility']})")

print(f"\n3. 🔍 Sai lệch Số Giấy Chứng Nhận (Cert No Discrepancies):")
print(f"   • Số lượng: {len(discrepancies['cert_no_mismatches'])} máy")
for m in discrepancies['cert_no_mismatches'][:5]:
    print(f"     - [{m['name']}] SN: {m['serial']} | DB: {m['db_cert']} vs Master: {m['master_cert']}")

print(f"\n4. 🔍 Thiết bị trong DB chưa được phân bổ Khoa/Phòng (Unassigned):")
print(f"   • Số lượng: {len(discrepancies['unassigned_facility_in_db'])} máy")
for m in discrepancies['unassigned_facility_in_db'][:5]:
    print(f"     - [ID {m['id']:04d}] {m['device_name']} (Model: {m['model']} | SN: {m['serial_no']})")

print(f"\n5. 🔍 Sai lệch Khoa Phòng giữa Master và DB (Facility Mismatches):")
print(f"   • Số lượng: {len(discrepancies['facility_mismatches'])} máy")
for m in discrepancies['facility_mismatches'][:5]:
    print(f"     - [{m['name']}] SN: {m['serial']} | DB: {m['db_fac']} vs Master: {m['master_fac']}")

conn.close()
