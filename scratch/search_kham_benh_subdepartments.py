import sqlite3
import openpyxl
import csv
import sys
from pathlib import Path
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8')

ocr_root = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712")

print("🔍 TRA CỨU TẤT CẢ CÁC PHÒNG KHÁM CHUYÊN KHOA THUỘC 'KHOA KHÁM BỆNH':\n" + "=" * 75)

# 1. Tra cứu trong Master Data.xltm
xltm_path = ocr_root / "Master Data.xltm"
xltm_clinics = set()
if xltm_path.exists():
    wb = openpyxl.load_workbook(xltm_path, data_only=True)
    ws2 = wb['2. Ban giao lap dat']
    for r in range(2, ws2.max_row + 1):
        dept = ws2.cell(r, 10).value
        room_no = ws2.cell(r, 11).value
        room_name = ws2.cell(r, 12).value
        if dept and 'khám' in str(dept).lower():
            xltm_clinics.add((str(dept).strip(), str(room_no or '').strip(), str(room_name or '').strip()))
            
    ws4 = wb['Dropdown']
    dropdown_depts = [ws4.cell(r, 5).value for r in range(2, ws4.max_row + 1) if ws4.cell(r, 5).value]
    print(f"📊 Danh mục Khoa/Phòng trong Dropdown của Master Data.xltm:")
    for d in dropdown_depts:
        print(f"   • {d}")

# 2. Tra cứu trong Master_kiem_dinh_TB.md
md_path = ocr_root / "md" / "05_KIEM DINH" / "pdf" / "Master_kiem_dinh_TB.md"
md_clinics = Counter()
if md_path.exists():
    with open(md_path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip().startswith('|') and 'Khoa /Phòng' not in line:
                cols = [c.strip() for c in line.strip().split('|')[1:-1]]
                if cols and 'KHÁM' in cols[0].upper():
                    md_clinics[cols[0]] += 1

print("\n📊 Các phân khoa / phòng khám thuộc KHÁM BỆNH trong 'Master_kiem_dinh_TB.md':")
for c_name, count in md_clinics.most_common():
    print(f"   • [{c_name}]: {count} thiết bị")

# 3. Tra cứu trong handover_master_enriched.csv
handover_csv = ocr_root / "03_BAN_GIAO_VA_NGHIEM_THU" / "_ocr_handover_assets" / "handover_master_enriched.csv"
handover_clinics = Counter()
if handover_csv.exists():
    with open(handover_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for r in reader:
            d = (r.get('department') or '').strip()
            if 'khám' in d.lower() or 'phòng' in d.lower():
                handover_clinics[d] += 1

print("\n📊 Các phân khoa / phòng khám trong 'handover_master_enriched.csv':")
for c_name, count in handover_clinics.most_common(15):
    print(f"   • [{c_name}]: {count} thiết bị")
