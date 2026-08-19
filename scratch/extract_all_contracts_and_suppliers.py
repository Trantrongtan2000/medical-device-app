import openpyxl
import sqlite3
import os
import sys
import re
from pathlib import Path
from collections import defaultdict, Counter

sys.stdout.reconfigure(encoding='utf-8')

ocr_root = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712")
g_bvq7 = Path(r"G:\BV QUẬN 7")
db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")

print("🔍 BẮT ĐẦU TRÍCH XUẤT TOÀN DIỆN TẤT CẢ HỢP ĐỒNG & NHÀ THẦU CUNG CẤP:\n" + "=" * 75)

all_contracts = {} # contract_no -> {supplier, sign_date, equipment_list: []}
all_suppliers = set()
device_contract_map = {} # serial_no / model / name -> {contract_no, supplier_name, handover_date}

# 1. Trích xuất từ Master Data.xltm Sheet 1 [1. Hop dong mua sam] & Sheet 2 [2. Ban giao lap dat]
xltm_path = ocr_root / "Master Data.xltm"
if xltm_path.exists():
    wb = openpyxl.load_workbook(xltm_path, data_only=True)
    
    # Sheet 1
    ws1 = wb['1. Hop dong mua sam']
    for r in range(2, ws1.max_row + 1):
        c_no = ws1.cell(r, 2).value
        c_date = ws1.cell(r, 3).value
        sup = ws1.cell(r, 4).value
        eq_name = ws1.cell(r, 5).value
        model = ws1.cell(r, 6).value
        qty = ws1.cell(r, 7).value
        mfg = ws1.cell(r, 8).value
        country = ws1.cell(r, 9).value
        h_date = ws1.cell(r, 11).value or ws1.cell(r, 10).value
        
        if c_no:
            c_str = str(c_no).strip()
            sup_str = str(sup or '').strip()
            if sup_str:
                all_suppliers.add(sup_str)
            if c_str not in all_contracts:
                all_contracts[c_str] = {
                    "contract_no": c_str,
                    "supplier_name": sup_str,
                    "date": str(c_date)[:10] if c_date else "",
                    "handover_date": str(h_date)[:10] if h_date else "",
                    "items": []
                }
            all_contracts[c_str]["items"].append({
                "name": str(eq_name or '').strip(),
                "model": str(model or '').strip(),
                "mfg": str(mfg or '').strip()
            })
            
    # Sheet 2
    ws2 = wb['2. Ban giao lap dat']
    for r in range(2, ws2.max_row + 1):
        c_no = ws2.cell(r, 2).value
        sup = ws2.cell(r, 3).value
        eq_name = ws2.cell(r, 4).value
        model = ws2.cell(r, 5).value
        sn = ws2.cell(r, 8).value
        dept = ws2.cell(r, 10).value
        
        c_str = str(c_no or '').strip()
        sup_str = str(sup or '').strip()
        sn_str = str(sn or '').strip().upper()
        
        if sup_str:
            all_suppliers.add(sup_str)
        if c_str and c_str not in all_contracts:
            all_contracts[c_str] = {"contract_no": c_str, "supplier_name": sup_str, "items": []}
            
        if sn_str and sn_str not in ['NONE', 'N/A', 'KHÔNG CÓ', '-', '']:
            device_contract_map[sn_str] = {
                "contract_no": c_str or "HĐMB-Q7-GENERAL",
                "supplier_name": sup_str or "Nhà thầu y tế",
                "department": str(dept or '').strip(),
                "model": str(model or '').strip(),
                "name": str(eq_name or '').strip()
            }
            
    # Sheet 4 Dropdown suppliers
    ws4 = wb['Dropdown']
    for r in range(2, ws4.max_row + 1):
        sup_dd = ws4.cell(r, 2).value
        if sup_dd:
            all_suppliers.add(str(sup_dd).strip())

print(f"📊 Từ Master Data.xltm: {len(all_contracts)} Hợp đồng, {len(all_suppliers)} Nhà thầu, {len(device_contract_map)} Máy có Serial")

# 2. Trích xuất từ cây thư mục PDF 02_HOP DONG MUA SAM tại G:\BV QUẬN 7 và OCR WORK
def parse_contracts_from_folder(base_p: Path):
    if not base_p.exists():
        return
    for dirpath, dirnames, filenames in os.walk(base_p):
        for f in filenames:
            fl = f.lower()
            if fl.endswith('.pdf'):
                # Extract potential supplier from folder name or file name
                rel_parts = Path(dirpath).relative_to(base_p).parts
                sup_candidate = rel_parts[0] if rel_parts else ""
                
                # Check for Contract patterns in filename
                # Patterns like HD 4005, HD 05525, HD 023, 12825/HĐMB, PO 25020152, HD TB01, etc.
                m_hd = re.search(r'(HĐ[^\._]+|HD[^\._]+|PO[^\._]+|\d{4,}[^\._]*HĐ[^\._]*|\d{5,})', f, re.IGNORECASE)
                if m_hd:
                    found_contract = m_hd.group(1).strip()
                    if len(found_contract) >= 4:
                        if found_contract not in all_contracts:
                            all_contracts[found_contract] = {
                                "contract_no": found_contract,
                                "supplier_name": sup_candidate if sup_candidate not in ['Biên bản bàn giao nội bộ', 'docs_raw', 'HOP_DONG_GOC'] else "Nhà cung cấp theo HĐ",
                                "items": []
                            }
                if sup_candidate and sup_candidate not in ['Biên bản bàn giao nội bộ', 'docs_raw', 'HOP_DONG_GOC', '2024', '2025', '2026', 'CHUNG']:
                    all_suppliers.add(sup_candidate)

parse_contracts_from_folder(ocr_root / "02_HOP_DONG_MUA_SAM")
parse_contracts_from_folder(g_bvq7 / "02_HOP DONG MUA SAM")

# 3. Trích xuất từ tất cả các file Markdown trong md/
md_root = ocr_root / "md"
if md_root.exists():
    for dirpath, dirnames, filenames in os.walk(md_root):
        for f in filenames:
            if f.endswith('.md'):
                fp = os.path.join(dirpath, f)
                try:
                    with open(fp, 'r', encoding='utf-8', errors='ignore') as file:
                        lines = [file.readline() for _ in range(30)]
                        c_no = ""
                        mfg = ""
                        giver = ""
                        sn = ""
                        model = ""
                        eq_name = ""
                        for l in lines:
                            if l.startswith('---') and lines.index(l) > 0:
                                break
                            if ':' in l:
                                k, v = l.split(':', 1)
                                k_s = k.strip().lower()
                                v_s = v.strip().strip('"').strip("'")
                                if k_s == 'contract_no' and v_s and v_s != 'None':
                                    c_no = v_s
                                elif k_s in ['manufacturer', 'party_giver'] and v_s and v_s != 'None':
                                    if k_s == 'party_giver' and ('cty' in v_s.lower() or 'công ty' in v_s.lower() or 'ct' in v_s.lower()):
                                        giver = v_s
                                    elif k_s == 'manufacturer':
                                        mfg = v_s
                                elif k_s == 'serial_no':
                                    sn = v_s.upper()
                                elif k_s == 'model':
                                    model = v_s
                                elif k_s == 'equipment_name':
                                    eq_name = v_s
                        if c_no:
                            if c_no not in all_contracts:
                                all_contracts[c_no] = {"contract_no": c_no, "supplier_name": giver or mfg or "Nhà thầu", "items": []}
                        if giver:
                            all_suppliers.add(giver)
                        if sn and sn not in ['NONE', 'N/A', '-', '']:
                            device_contract_map[sn] = {
                                "contract_no": c_no or "HĐMB-Q7-GENERAL",
                                "supplier_name": giver or mfg or "Nhà thầu y tế",
                                "model": model,
                                "name": eq_name
                            }
                except Exception:
                    continue

print("\n" + "=" * 75)
print(f"🎉 TỔNG HỢP TOÀN BỘ KHO DỮ LIỆU:")
print(f"  • Tổng số Hợp Đồng / Gói Thầu Mua Sắm: {len(all_contracts):,} Hợp Đồng")
print(f"  • Tổng số Đơn Vị / Nhà Thầu Cung Cấp: {len(all_suppliers):,} Nhà Thầu")
print(f"  • Tổng số Thiết Bị có liên kết Serial chuẩn xác: {len(device_contract_map):,} máy")
print("=" * 75)

print("\n📑 DANH SÁCH 35 NHÀ THẦU CUNG CẤP Y TẾ TIÊU BIỂU:")
for idx, sup in enumerate(sorted(all_suppliers)[:35], 1):
    print(f"  {idx:02d}. {sup}")

print("\n📑 DANH SÁCH 35 HỢP ĐỒNG / GÓI MUA SẮM TIÊU BIỂU:")
for idx, (c_no, c_data) in enumerate(sorted(all_contracts.items())[:35], 1):
    print(f"  {idx:02d}. [HĐ: {c_no}] — {c_data.get('supplier_name') or 'Nhà thầu y tế'}")
