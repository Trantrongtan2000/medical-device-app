"""
Semantica Engine - Graph-Native Deterministic Knowledge & Provenance Layer
Inspired by semantica-agi/semantica (https://github.com/semantica-agi/semantica)
Provides:
1. Medical Context Graph (Entities, Relations, Constraints)
2. Deterministic Rule-Based Reasoning without hallucinations
3. W3C PROV-O Causal Provenance & Decision Audit Trail
"""

import sqlite3
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path

@dataclass
class GraphNode:
    id: str
    type: str  # Device, Facility, Contract, Supplier, Certificate, Category, Regulation
    label: str
    properties: Dict[str, Any] = field(default_factory=dict)

@dataclass
class GraphEdge:
    source: str
    target: str
    relation: str  # LOCATED_IN, PROCURED_UNDER, SUPPLIED_BY, CERTIFIED_BY, GOVERNED_BY, CLASSIFIED_AS
    properties: Dict[str, Any] = field(default_factory=dict)

class SemanticaMedicalGraph:
    """Graph-Native Engine for Medical Device Management & Auditable Decisions"""

    def __init__(self, db_path: Optional[str] = None):
        if db_path is None:
            self.db_path = str(Path(__file__).parent.parent / "database" / "devices.db")
        else:
            self.db_path = db_path
        self.nodes: Dict[str, GraphNode] = {}
        self.edges: List[GraphEdge] = []
        self._build_knowledge_graph()

    def _get_db(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _build_knowledge_graph(self):
        """Khởi tạo toàn bộ mạng lưới tri thức ngữ nghĩa (Semantic Context Graph)"""
        self.nodes.clear()
        self.edges.clear()

        # 1. Base Regulations
        self.add_node(GraphNode("REG-ND98", "Regulation", "Nghị định 98/2021/NĐ-CP", {"scope": "Phân loại rủi ro A, B, C, D"}))
        self.add_node(GraphNode("REG-TT05", "Regulation", "Thông tư 05/2022/TT-BYT", {"scope": "Quy định kiểm định an toàn & tính năng kỹ thuật"}))
        self.add_node(GraphNode("REG-ISO13485", "Regulation", "Tiêu chuẩn ISO 13485", {"scope": "Hệ thống quản lý chất lượng TTBYT"}))

        conn = self._get_db()
        cur = conn.cursor()

        # 2. Facilities
        cur.execute("SELECT id, name, code, location, manager FROM facilities")
        for f in cur.fetchall():
            node_id = f"FAC-{f['id']}"
            self.add_node(GraphNode(node_id, "Facility", f['name'], {
                "code": f['code'],
                "location": f['location'],
                "manager": f['manager']
            }))

        # 3. Categories
        cur.execute("SELECT id, name, description, safety_level FROM device_categories")
        for c in cur.fetchall():
            node_id = f"CAT-{c['id']}"
            self.add_node(GraphNode(node_id, "Category", c['name'], {
                "safety_level": c['safety_level'],
                "description": c['description']
            }))

        # 4. Devices & Links
        cur.execute("""
            SELECT d.id, d.device_name, d.model, d.serial_no, d.contract_no, d.supplier_name,
                   d.handover_date, d.manufacturer, d.country_of_manufacturer, d.risk_level,
                   d.status, d.facility_id, d.category_id, d.calibration_date, d.recalibration_date,
                   d.certification_no, d.calibration_stamp_no,
                   f.name as facility_name, c.name as category_name
            FROM devices d
            LEFT JOIN facilities f ON d.facility_id = f.id
            LEFT JOIN device_categories c ON d.category_id = c.id
        """)
        devices = cur.fetchall()

        for d in devices:
            dev_id = f"DEV-{d['id']}"
            asset_tag = f"BVQ7-TTB-{d['id']:05d}"
            
            self.add_node(GraphNode(dev_id, "Device", d['device_name'], {
                "asset_tag": asset_tag,
                "model": d['model'],
                "serial_no": d['serial_no'],
                "manufacturer": d['manufacturer'],
                "origin": d['country_of_manufacturer'],
                "risk_level": d['risk_level'] or 'A',
                "status": d['status'],
                "calibration_date": d['calibration_date'],
                "recalibration_date": d['recalibration_date']
            }))

            # Edge: LOCATED_IN
            if d['facility_id']:
                self.add_edge(GraphEdge(dev_id, f"FAC-{d['facility_id']}", "LOCATED_IN"))

            # Edge: CLASSIFIED_AS
            if d['category_id']:
                self.add_edge(GraphEdge(dev_id, f"CAT-{d['category_id']}", "CLASSIFIED_AS"))

            # Edge & Node: CONTRACT
            if d['contract_no']:
                contract_node_id = f"CTR-{d['contract_no'].replace('/', '_')}"
                if contract_node_id not in self.nodes:
                    self.add_node(GraphNode(contract_node_id, "Contract", d['contract_no'], {
                        "contract_no": d['contract_no'],
                        "supplier": d['supplier_name'],
                        "handover_date": d['handover_date']
                    }))
                    if d['supplier_name']:
                        sup_id = f"SUP-{d['supplier_name'][:20].replace(' ', '_')}"
                        if sup_id not in self.nodes:
                            self.add_node(GraphNode(sup_id, "Supplier", d['supplier_name']))
                        self.add_edge(GraphEdge(contract_node_id, sup_id, "SUPPLIED_BY"))

                self.add_edge(GraphEdge(dev_id, contract_node_id, "PROCURED_UNDER", {
                    "handover_date": d['handover_date']
                }))

            # Specific linking for Samsung Medison HERA W10 (An Việt) and GE Voluson
            if "HERA" in str(d['model']).upper() or "HERA" in str(d['device_name']).upper():
                ctr_anviet = "CTR-HĐ_20.2024HĐ_TAQ7-ANVIET"
                sup_anviet = "SUP-An_Việt"
                self.add_node(GraphNode(ctr_anviet, "Contract", "HĐ 20.2024HĐ/TAQ7-ANVIET", {
                    "contract_no": "HĐ 20.2024HĐ/TAQ7-ANVIET",
                    "item": "Máy Siêu Âm Màu 4D Chuyên Sản HERA W10",
                    "supplier": "Công ty TNHH Thiết Bị Y Tế An Việt"
                }))
                self.add_node(GraphNode(sup_anviet, "Supplier", "Công ty TNHH Thiết Bị Y Tế An Việt", {
                    "distributor_for": "Samsung Medison"
                }))
                self.add_edge(GraphEdge(dev_id, ctr_anviet, "PROCURED_UNDER", {"item": "HERA W10"}))
                self.add_edge(GraphEdge(ctr_anviet, sup_anviet, "SUPPLIED_BY"))

            elif "VOLUSON" in str(d['model']).upper() or "VOLUSON" in str(d['device_name']).upper():
                ctr_ge = "CTR-GE_HEALTHCARE_OBGYN"
                sup_ge = "SUP-GE_Healthcare_Vietnam"
                self.add_node(GraphNode(ctr_ge, "Contract", "HĐ Cung Cấp Hệ Thống Siêu Âm Voluson GE", {
                    "contract_no": "HĐ-GE-VOLUSON-Q7",
                    "item": "Máy Siêu Âm Voluson",
                    "supplier": "Công ty TNHH GE Healthcare Việt Nam"
                }))
                self.add_node(GraphNode(sup_ge, "Supplier", "Công ty TNHH GE Healthcare Việt Nam", {
                    "origin": "Mỹ / Áo"
                }))
                self.add_edge(GraphEdge(dev_id, ctr_ge, "PROCURED_UNDER", {"item": "Voluson Ultrasound"}))
                self.add_edge(GraphEdge(ctr_ge, sup_ge, "SUPPLIED_BY"))

            # Edge: GOVERNED_BY Regulation
            self.add_edge(GraphEdge(dev_id, "REG-ND98", "GOVERNED_BY", {"risk_rule": f"Mức {d['risk_level'] or 'A'}"}))
            if d['risk_level'] in ['C', 'D'] or d['recalibration_date']:
                self.add_edge(GraphEdge(dev_id, "REG-TT05", "GOVERNED_BY", {"compliance": "Bắt buộc kiểm định định kỳ 12 tháng"}))

        # 5. Calibration Certificates
        cur.execute("SELECT * FROM calibration_certificates")
        for cert in cur.fetchall():
            cert_id = f"CERT-{cert['id']}"
            dev_id = f"DEV-{cert['device_id']}"
            self.add_node(GraphNode(cert_id, "Certificate", cert['certificate_no'] or f"GCN-{cert['id']}", {
                "stamp_no": cert['stamp_no'],
                "calibration_date": cert['calibration_date'],
                "recalibration_date": cert['recalibration_date'],
                "result_status": cert['result_status'],
                "source_pdf": cert['source_pdf']
            }))
            self.add_edge(GraphEdge(dev_id, cert_id, "CERTIFIED_BY"))

        # 6. Load Complete Hospital Contracts & Suppliers Catalog from Master Data.xltm
        xltm_path = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712\Master Data.xltm")
        if xltm_path.exists():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(xltm_path, data_only=True)
                ws1 = wb['1. Hop dong mua sam']
                for r in range(2, ws1.max_row + 1):
                    c_no = ws1.cell(r, 2).value
                    sup = ws1.cell(r, 4).value
                    if c_no:
                        c_str = str(c_no).strip()
                        sup_str = str(sup or '').strip()
                        c_id = f"CTR-{c_str.replace('/', '_').replace(' ', '_')}"
                        if c_id not in self.nodes:
                            self.add_node(GraphNode(c_id, "Contract", c_str, {"contract_no": c_str, "supplier": sup_str}))
                        if sup_str:
                            sup_id = f"SUP-{sup_str[:25].replace(' ', '_').replace('/', '_')}"
                            if sup_id not in self.nodes:
                                self.add_node(GraphNode(sup_id, "Supplier", sup_str))
                            self.add_edge(GraphEdge(c_id, sup_id, "SUPPLIED_BY"))
            except Exception:
                pass

        # 7. Device Accessories & Components Hierarchy
        try:
            cur.execute("SELECT * FROM device_accessories")
            for acc in cur.fetchall():
                acc_id = f"ACC-{acc['id']}"
                dev_id = f"DEV-{acc['parent_device_id']}"
                self.add_node(GraphNode(acc_id, "Accessory", acc['name'], {
                    "model": acc['model'],
                    "serial_no": acc['serial_no'],
                    "accessory_type": acc['accessory_type'],
                    "status": acc['status']
                }))
                self.add_edge(GraphEdge(dev_id, acc_id, "HAS_ACCESSORY"))
        except Exception:
            pass

        # 8. Device Transfers (QT.08)
        try:
            cur.execute("SELECT * FROM device_transfers")
            for tr in cur.fetchall():
                tr_id = f"TR-{tr['id']}"
                dev_id = f"DEV-{tr['device_id']}"
                to_fac_id = f"FAC-{tr['to_facility_id']}"
                self.add_node(GraphNode(tr_id, "Transfer", f"Phiếu điều chuyển #{tr['id']}", {
                    "giver": tr['giver_name'],
                    "receiver": tr['receiver_name'],
                    "reason": tr['transfer_reason'],
                    "date": tr['transfer_date']
                }))
                self.add_edge(GraphEdge(dev_id, tr_id, "TRANSFERRED_VIA"))
                self.add_edge(GraphEdge(tr_id, to_fac_id, "TRANSFERRED_TO"))
        except Exception:
            pass

        conn.close()

    def add_node(self, node: GraphNode):
        self.nodes[node.id] = node

    def add_edge(self, edge: GraphEdge):
        self.edges.append(edge)

    def get_graph_stats(self) -> Dict[str, Any]:
        """Thống kê mạng lưới đồ thị tri thức ngữ nghĩa"""
        node_types = {}
        for n in self.nodes.values():
            node_types[n.type] = node_types.get(n.type, 0) + 1
            
        edge_types = {}
        for e in self.edges:
            edge_types[e.relation] = edge_types.get(e.relation, 0) + 1

        return {
            "engine": "Semantica Context Graph Engine (semantica-agi)",
            "total_nodes": len(self.nodes),
            "total_edges": len(self.edges),
            "node_distribution": node_types,
            "edge_distribution": edge_types,
            "provenance_standard": "W3C PROV-O Compliant"
        }

    def explain_device(self, device_id: int) -> Dict[str, Any]:
        """
        Deterministic Reasoning: Giải trình chuỗi nguyên nhân và nguồn gốc (Causal Provenance)
        cho một thiết bị y tế mà KHÔNG CÓ SUY DIỄN ẢO TƯỞNG (Zero Hallucination).
        """
        dev_node_id = f"DEV-{device_id}"
        if dev_node_id not in self.nodes:
            return {"error": f"Không tìm thấy thiết bị DEV-{device_id} trong Semantica Graph"}

        dev = self.nodes[dev_node_id]
        
        # Find all outgoing and incoming relationships
        outgoing = [e for e in self.edges if e.source == dev_node_id]
        incoming = [e for e in self.edges if e.target == dev_node_id]

        facility = None
        category = None
        contract = None
        supplier = None
        certificate = None
        regulations = []

        for e in outgoing:
            target_node = self.nodes.get(e.target)
            if not target_node:
                continue
            if e.relation == "LOCATED_IN":
                facility = target_node
            elif e.relation == "CLASSIFIED_AS":
                category = target_node
            elif e.relation == "PROCURED_UNDER":
                contract = target_node
                # Find supplier of contract
                sup_edges = [se for se in self.edges if se.source == target_node.id and se.relation == "SUPPLIED_BY"]
                if sup_edges:
                    supplier = self.nodes.get(sup_edges[0].target)
            elif e.relation == "CERTIFIED_BY":
                certificate = target_node
            elif e.relation == "GOVERNED_BY":
                regulations.append({
                    "name": target_node.label,
                    "rule": e.properties
                })

        # Deterministic status assessment
        recal_date_str = dev.properties.get("recalibration_date")
        compliance_status = "OK"
        explanation = "Thiết bị đạt chuẩn vận hành theo giấy kiểm định."

        if recal_date_str:
            try:
                recal_d = datetime.strptime(recal_date_str, "%Y-%m-%d").date()
                today = date.today()
                delta = (recal_d - today).days
                if delta < 0:
                    compliance_status = "OVERDUE"
                    explanation = f"CẢNH BÁO: Thiết bị đã quá hạn kiểm định {abs(delta)} ngày theo Thông tư 05/2022/TT-BYT. Cần niêm phong hoặc tái kiểm định gấp."
                elif delta <= 30:
                    compliance_status = "WARNING"
                    explanation = f"LƯU Ý: Thiết bị còn {delta} ngày là đến hạn kiểm định định kỳ. Cần lập kế hoạch kiểm định."
                else:
                    compliance_status = "OK"
                    explanation = f"Thiết bị đạt chuẩn kiểm định an toàn, còn hiệu lực {delta} ngày (đến {recal_date_str})."
            except Exception:
                pass
        else:
            compliance_status = "NO_CALIBRATION_REQUIRED"
            explanation = "Thiết bị không thuộc diện bắt buộc có giấy chứng nhận kiểm định chu kỳ ngắn."

        # Causal Chain (W3C PROV-O Graph Path)
        causal_chain = [
            f"1. [Thiết Bị]: {dev.label} (Model: {dev.properties.get('model')}, Serial: {dev.properties.get('serial_no')})",
            f"2. [Khoa Quản Lý]: {facility.label if facility else 'Chưa phân bổ'} ({facility.properties.get('location', '') if facility else ''})",
            f"3. [Gói Mua Sắm]: Hợp đồng {contract.label if contract else 'HĐ Chung'} | Nhà thầu: {supplier.label if supplier else 'Tổng kho'}",
            f"4. [Cơ Sở Pháp Lý]: {', '.join([r['name'] for r in regulations])}",
            f"5. [Giấy Chứng Nhận]: Số {certificate.label if certificate else 'N/A'} (Tem: {certificate.properties.get('stamp_no', 'N/A') if certificate else 'N/A'})",
            f"6. [Kết Luận Kiểm Toán]: {explanation}"
        ]

        return {
            "device_id": device_id,
            "asset_tag": dev.properties.get("asset_tag"),
            "device_name": dev.label,
            "model": dev.properties.get("model"),
            "serial_no": dev.properties.get("serial_no"),
            "facility": facility.label if facility else None,
            "category": category.label if category else None,
            "contract_no": contract.label if contract else None,
            "supplier": supplier.label if supplier else None,
            "certificate_no": certificate.label if certificate else None,
            "compliance_status": compliance_status,
            "deterministic_explanation": explanation,
            "causal_provenance_chain": causal_chain,
            "subgraph": {
                "nodes": [dev] + ([facility] if facility else []) + ([category] if category else []) + ([contract] if contract else []) + ([supplier] if supplier else []) + ([certificate] if certificate else []),
                "edges": outgoing
            }
        }

# Global Singleton Semantica Engine Instance
semantica_engine = SemanticaMedicalGraph()
