import sqlite3
import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent.parent))

db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")
excel_path = Path(r"C:\Users\tantt\Downloads\CĐHA - Danh sách máy siêu âm 12-08-2026 NEW Q7.xlsx")

print("🏥 ĐỒNG BỘ DỮ LIỆU CHUYÊN SÂU MÁY SIÊU ÂM & THIẾT BỊ HÌNH ẢNH CĐHA Q7:\n" + "=" * 75)

conn = sqlite3.connect(db_path)
cur = conn.cursor()

wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Đọc và nạp 24 Máy Siêu Âm & Đầu Dò (Sheet SÂ)
ws_sa = wb['SÂ']
sa_machines = []
current_m = None

for r in range(3, ws_sa.max_row + 1):
    stt = ws_sa.cell(r, 1).value
    room = ws_sa.cell(r, 2).value
    m_name = ws_sa.cell(r, 3).value
    qty = ws_sa.cell(r, 4).value
    sn = ws_sa.cell(r, 5).value
    probe_name = ws_sa.cell(r, 6).value
    probe_qty = ws_sa.cell(r, 7).value
    probe_sn = ws_sa.cell(r, 8).value
    year = ws_sa.cell(r, 9).value
    notes = ws_sa.cell(r, 10).value

    if stt and (room or m_name):
        current_m = {
            "room": str(room or '').strip(),
            "name": str(m_name or '').strip(),
            "serial_no": str(sn or '').strip().replace('.0', ''),
            "year": str(year or '').strip().replace('.0', ''),
            "notes": str(notes or '').strip(),
            "probes": [],
            "ups": None
        }
        sa_machines.append(current_m)

    if current_m:
        if m_name and "UPS" in str(m_name).upper():
            current_m["ups"] = str(sn or '').strip().replace('.0', '')
        if probe_name:
            current_m["probes"].append({
                "probe_name": str(probe_name).strip(),
                "probe_sn": str(probe_sn or '').strip().replace('.0', '')
            })

probes_inserted = 0
machines_updated = 0

for m in sa_machines:
    sn = m['serial_no']
    # Match device in devices table by serial or name
    cur.execute("SELECT id FROM devices WHERE UPPER(serial_no) = ? OR (device_name LIKE ? AND model LIKE ?)", 
                (sn.upper(), f"%{m['name'][:10]}%", f"%{m['name'][-10:]}%"))
    row = cur.fetchone()
    
    dev_id = None
    if row:
        dev_id = row[0]
        machines_updated += 1
    else:
        # Insert if not exists in CĐHA (facility_id = 3)
        cur.execute("""
            INSERT INTO devices (device_name, model, serial_no, manufacturer, facility_id, category_id, risk_level, status)
            VALUES (?, ?, ?, 'Hãng CĐHA', 3, 3, 'C', 'IN_SERVICE')
        """, (f"Máy siêu âm {m['name']}", m['name'], sn or f"SA-Q7-{m['room']}"))
        dev_id = cur.lastrowid
        machines_updated += 1

    if dev_id:
        # Clear old accessories for this device
        cur.execute("DELETE FROM device_accessories WHERE parent_device_id = ?", (dev_id,))
        for p in m['probes']:
            cur.execute("""
                INSERT INTO device_accessories (parent_device_id, name, model, serial_no, accessory_type, status, notes)
                VALUES (?, ?, ?, ?, 'Probe', 'Sẵn sàng sử dụng', ?)
            """, (dev_id, p['probe_name'], m['name'], p['probe_sn'], f"Phòng {m['room']}"))
            probes_inserted += 1
            
        if m['ups']:
            cur.execute("""
                INSERT INTO device_accessories (parent_device_id, name, model, serial_no, accessory_type, status, notes)
                VALUES (?, 'Bộ lưu điện UPS chuyên dụng', 'UPS-SA', ?, 'Battery', 'Sẵn sàng sử dụng', ?)
            """, (dev_id, m['ups'], f"Phòng {m['room']}"))
            probes_inserted += 1

print(f"✅ Đã đồng bộ {machines_updated} Máy Siêu Âm CĐHA và nạp {probes_inserted} Đầu Dò / Phụ Kiện UPS vào CSDL!")

# 2. Đọc và nạp 21 Hệ Thống Kỹ Thuật Cao (Sheet XQCTMRINAĐLX)
ws_xq = wb['XQCTMRINAĐLX']
xq_updated = 0

for r in range(3, ws_xq.max_row + 1):
    stt = ws_xq.cell(r, 1).value
    room = ws_xq.cell(r, 2).value
    m_name = ws_xq.cell(r, 3).value
    sn = ws_xq.cell(r, 5).value
    year = ws_xq.cell(r, 6).value
    notes = ws_xq.cell(r, 7).value
    
    if stt and (room or m_name):
        room_str = str(room or '').strip().replace('.0', '')
        name_str = str(m_name or '').strip()
        sn_str = str(sn or '').strip().replace('.0', '')
        
        # Match device
        cur.execute("SELECT id FROM devices WHERE UPPER(serial_no) = ? OR (device_name LIKE ? AND facility_id = 3)",
                    (sn_str.upper() if sn_str else "NONE", f"%{name_str[:15]}%"))
        row = cur.fetchone()
        if row:
            xq_updated += 1
        else:
            risk = 'D' if any(k in name_str.upper() for k in ['CHT', 'MRI', 'CT', 'SOMATOM', 'REVOLUTION']) else 'C'
            cur.execute("""
                INSERT INTO devices (device_name, model, serial_no, manufacturer, facility_id, category_id, risk_level, status)
                VALUES (?, ?, ?, 'Hãng thiết bị CĐHA', 3, 3, ?, 'IN_SERVICE')
            """, (name_str, name_str, sn_str or f"CĐHA-Q7-P{room_str}", risk))
            xq_updated += 1

print(f"✅ Đã đồng bộ {xq_updated} Hệ thống Kỹ thuật cao (MRI 3T/1.5T, CT đa lát cắt, X-Quang, Nhũ ảnh, Đo loãng xương)!")

conn.commit()

# Rebuild Semantica Engine
from app.semantica_engine import semantica_engine
semantica_engine._build_knowledge_graph()
stats = semantica_engine.get_graph_stats()

print(f"\n🕸️ SEMANTICA AGI KNOWLEDGE GRAPH SAU KHI ĐỒNG BỘ CĐHA:")
print(f"  • Tổng số Thực Thể (Nodes): {stats['total_nodes']:,} nodes")
print(f"  • Tổng số Mối Quan Hệ (Edges): {stats['total_edges']:,} edges")
print(f"  • Phân bổ Nodes: {stats['node_distribution']}")
print(f"  • Phân bổ Edges: {stats['edge_distribution']}")

conn.close()
