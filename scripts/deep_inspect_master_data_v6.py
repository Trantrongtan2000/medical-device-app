import openpyxl
import sys
from pathlib import Path
from collections import defaultdict, Counter

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\tantt\Downloads\MasterData_V6_V1.0 -USERFORM MODEL_439_MERGE_MUNUAL.xlsm")
print("="*90)
print(f"📖 ĐỌC CHI TIẾT MASTER DATA TỪ: {excel_path.name}")
print("="*90)

wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Inspect Bangiao sheet
ws_bg = wb["Bangiao"]
bg_rows = list(ws_bg.iter_rows(values_only=True))
bg_headers = bg_rows[0]
print(f"\n📌 SHEET 'Bangiao': {len(bg_rows)-1} bản ghi tài sản")
print("Cột:", bg_headers)

# Sample rows
print("\n--- 10 DÒNG ĐẦU TIÊN CỦA BANGIAO ---")
for r in bg_rows[1:11]:
    asset_id, proc_id, contract_no, supplier, name, model, mfg, origin, sn, ptype, dept = r[:11]
    print(f"  • [{asset_id}] {name} | Model: {model} | S/N: {sn} | Khoa: {dept}")
    print(f"    -> HĐ: {contract_no} | NCC: {supplier} | Hãng: {mfg} ({origin})")

# Count by Supplier in Bangiao
supplier_counts = Counter(r[3] for r in bg_rows[1:] if r[3])
print(f"\n📊 TỔNG HỢP {len(supplier_counts)} NHÀ CUNG CẤP TRONG SHEET 'Bangiao':")
for sup, cnt in supplier_counts.most_common():
    print(f"  • {sup}: {cnt} tài sản")

# Count by Contract in Bangiao
contract_counts = Counter(r[2] for r in bg_rows[1:] if r[2])
print(f"\n📊 TỔNG HỢP {len(contract_counts)} HỢP ĐỒNG TRONG SHEET 'Bangiao':")
for c, cnt in contract_counts.most_common()[:25]:
    print(f"  • HĐ: {c}: {cnt} tài sản")

# Count by Department in Bangiao
dept_counts = Counter(r[10] for r in bg_rows[1:] if r[10])
print(f"\n📊 TỔNG HỢP KHOA/PHÒNG TRONG SHEET 'Bangiao':")
for d, cnt in dept_counts.most_common():
    print(f"  • Khoa/Phòng: {d}: {cnt} tài sản")

# 2. Inspect Hopdongmuasam sheet
ws_hd = wb["Hopdongmuasam"]
hd_rows = list(ws_hd.iter_rows(values_only=True))
print(f"\n📌 SHEET 'Hopdongmuasam': {len(hd_rows)-1} dòng hợp đồng")

wb.close()
