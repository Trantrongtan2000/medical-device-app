import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\tantt\Downloads\MasterData_V6_V1.0 -USERFORM MODEL_439_MERGE_MUNUAL.xlsm")
print(f"Opening Excel file: {excel_path.name} ({excel_path.stat().st_size / 1024:.1f} KB)")

wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*40}\n📊 SHEET: {sheet_name}\n{'='*40}")
    
    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows: {len(rows)}")
    
    if len(rows) > 0:
        print("Header row (Row 1):", rows[0][:20])
    if len(rows) > 1:
        print("Row 2:", rows[1][:20])
    if len(rows) > 2:
        print("Row 3:", rows[2][:20])
    if len(rows) > 3:
        print("Row 4:", rows[3][:20])

wb.close()
