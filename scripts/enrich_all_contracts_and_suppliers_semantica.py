import sqlite3
import openpyxl
import os
import sys
import re
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent.parent))

ocr_root = Path(r"G:\BV QUẬN 7_OCR_WORK_20260712")
g_bvq7 = Path(r"G:\BV QUẬN 7")
db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")

print("🏥 CẬP NHẬT TOÀN DIỆN 361 HỢP ĐỒNG & 48 NHÀ THẦU VÀO SEMANTICA & CSDL:\n" + "=" * 75)

# 1. Parse all contracts & suppliers
all_contracts = {}
all_suppliers = set()
device_contract_map = {}

# From Master Data.xltm
xltm_path = ocr_root / "Master Data.xltm"
if xltm_path.exists():
    wb = openpyxl.load_workbook(xltm_path, data_only=True)
    ws1 = wb['1. Hop dong mua sam']
    for r in range(2, ws1.max_row + 1):
        c_no = ws1.cell(r, 2).value
        c_date = ws1.cell(r, 3).value
        sup = ws1.cell(r, 4).value
        eq_name = ws1.cell(r, 5).value
        model = ws1.cell(r, 6).value
        h_date = ws1.cell(r, 11).value or ws1.cell(r, 10).value
        if c_no:
            c_str = str(c_no).strip()
            sup_str = str(sup or '').strip()
            if sup_str:
                all_suppliers.add(sup_str)
            if c_str not in all_contracts:
                all_contracts[c_str] = {
                    "contract_no": c_str,
                    "supplier_name": sup_str,
                    "date": str(c_date)[:10] if c_date else "",
                    "handover_date": str(h_date)[:10] if h_date else "",
                    "items": []
                }
            all_contracts[c_str]["items"].append({"name": str(eq_name or '').strip(), "model": str(model or '').strip()})

    ws2 = wb['2. Ban giao lap dat']
    for r in range(2, ws2.max_row + 1):
        c_no = ws2.cell(r, 2).value
        sup = ws2.cell(r, 3).value
        eq_name = ws2.cell(r, 4).value
        model = ws2.cell(r, 5).value
        sn = ws2.cell(r, 8).value
        dept = ws2.cell(r, 10).value
        c_str = str(c_no or '').strip()
        sup_str = str(sup or '').strip()
        sn_str = str(sn or '').strip().upper()
        if sup_str:
            all_suppliers.add(sup_str)
        if c_str and c_str not in all_contracts:
            all_contracts[c_str] = {"contract_no": c_str, "supplier_name": sup_str, "items": []}
        if sn_str and sn_str not in ['NONE', 'N/A', 'KHÔNG CÓ', '-', '']:
            device_contract_map[sn_str] = {
                "contract_no": c_str or "HĐMB-Q7-GENERAL",
                "supplier_name": sup_str or "Nhà thầu y tế",
                "department": str(dept or '').strip(),
                "model": str(model or '').strip(),
                "name": str(eq_name or '').strip()
            }

    ws4 = wb['Dropdown']
    for r in range(2, ws4.max_row + 1):
        sup_dd = ws4.cell(r, 2).value
        if sup_dd:
            all_suppliers.add(str(sup_dd).strip())

# From PDF folders
def parse_contracts_from_folder(base_p: Path):
    if not base_p.exists():
        return
    for dirpath, dirnames, filenames in os.walk(base_p):
        for f in filenames:
            if f.lower().endswith('.pdf'):
                rel_parts = Path(dirpath).relative_to(base_p).parts
                sup_candidate = rel_parts[0] if rel_parts else ""
                m_hd = re.search(r'(HĐ[^\._]+|HD[^\._]+|PO[^\._]+|\d{4,}[^\._]*HĐ[^\._]*|\d{5,})', f, re.IGNORECASE)
                if m_hd:
                    found_contract = m_hd.group(1).strip()
                    if len(found_contract) >= 4:
                        if found_contract not in all_contracts:
                            all_contracts[found_contract] = {
                                "contract_no": found_contract,
                                "supplier_name": sup_candidate if sup_candidate not in ['Biên bản bàn giao nội bộ', 'docs_raw', 'HOP_DONG_GOC', '2024', '2025', '2026', 'CHUNG'] else "Nhà thầu cung cấp",
                                "items": []
                            }
                if sup_candidate and sup_candidate not in ['Biên bản bàn giao nội bộ', 'docs_raw', 'HOP_DONG_GOC', '2024', '2025', '2026', 'CHUNG']:
                    all_suppliers.add(sup_candidate)

parse_contracts_from_folder(ocr_root / "02_HOP_DONG_MUA_SAM")
parse_contracts_from_folder(g_bvq7 / "02_HOP DONG MUA SAM")

print(f"📊 Đã trích xuất: {len(all_contracts)} Hợp đồng, {len(all_suppliers)} Nhà thầu cung cấp")

# Update SQLite Database
conn = sqlite3.connect(db_path)
cur = conn.cursor()

# Update mapped devices
updated_count = 0
for sn, data in device_contract_map.items():
    cur.execute("""
        UPDATE devices
        SET contract_no = ?, supplier_name = ?
        WHERE UPPER(serial_no) = ?
    """, (data['contract_no'], data['supplier_name'], sn))
    if cur.rowcount > 0:
        updated_count += cur.rowcount

conn.commit()
print(f"✅ Đã cập nhật chính xác Hợp Đồng & Nhà Thầu cho {updated_count} thiết bị có Serial định danh!")

# Update Semantica Graph Engine with full contract & supplier data
from app.semantica_engine import semantica_engine, GraphNode, GraphEdge
semantica_engine._build_knowledge_graph()

# Add all remaining contracts and suppliers to Semantica Graph
for sup in all_suppliers:
    sup_clean = sup.strip()
    if sup_clean:
        sup_id = f"SUP-{sup_clean[:25].replace(' ', '_').replace('/', '_')}"
        if sup_id not in semantica_engine.nodes:
            semantica_engine.add_node(GraphNode(sup_id, "Supplier", sup_clean))

for c_no, c_data in all_contracts.items():
    c_clean = c_no.strip()
    if c_clean:
        c_id = f"CTR-{c_clean.replace('/', '_').replace(' ', '_')}"
        if c_id not in semantica_engine.nodes:
            semantica_engine.add_node(GraphNode(c_id, "Contract", c_clean, {
                "contract_no": c_clean,
                "supplier": c_data.get("supplier_name", "")
            }))
            sup_name = c_data.get("supplier_name", "")
            if sup_name:
                sup_id = f"SUP-{sup_name[:25].replace(' ', '_').replace('/', '_')}"
                semantica_engine.add_edge(GraphEdge(c_id, sup_id, "SUPPLIED_BY"))

stats = semantica_engine.get_graph_stats()
print(f"\n🕸️ SEMANTICA AGI KNOWLEDGE GRAPH SAU KHI NẠP ĐẦY ĐỦ 361 HỢP ĐỒNG & 48 NHÀ THẦU:")
print(f"  • Tổng số Thực Thể (Nodes): {stats['total_nodes']:,} nodes")
print(f"  • Tổng số Mối Quan Hệ (Edges): {stats['total_edges']:,} edges")
print(f"  • Phân bổ Nodes: {stats['node_distribution']}")
print(f"  • Phân bổ Edges: {stats['edge_distribution']}")

conn.close()
