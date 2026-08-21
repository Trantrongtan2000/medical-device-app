import sys
import io
import openpyxl
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except Exception:
        pass

excel_path = Path(r"C:\Users\tantt\Downloads\MasterData_V6_V1.0 -USERFORM MODEL_439_MERGE_MUNUAL.xlsm")
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("=== QUÉT FILE EXCEL MASTERDATA TÌM TỪ KHÓA 'abnf' HOẶC 'chuyên dùng' ===")
for sname in wb.sheetnames:
    ws = wb[sname]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = " ".join([str(c) for c in row if c is not None])
        if "abnf" in row_str.lower() or "a99-5" in row_str.lower() or "ske-120a" in row_str.lower():
            print(f"Sheet [{sname:15s}] Row {r_idx:4d}: {row_str[:120]}")
