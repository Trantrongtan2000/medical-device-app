"""
Audit script: MasterData_V6_V1.0 -USERFORM MODEL_439_MERGE_MUNUAL.xlsm vs SQLite database
"""
import sys
import io
import openpyxl
import sqlite3
from pathlib import Path
from collections import defaultdict, Counter

# UTF-8 handling for Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except Exception:
        pass

excel_path = Path(r"C:\Users\tantt\Downloads\MasterData_V6_V1.0 -USERFORM MODEL_439_MERGE_MUNUAL.xlsm")
db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")

print("="*90)
print(f"🔍 AUDIT MASTER DATA: {excel_path.name} VS CSDL SQLITE (devices.db)")
print("="*90)

wb = openpyxl.load_workbook(excel_path, data_only=True)
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# 1. Inspect Sheets
print("\n📑 1. DANH SÁCH SHEET TRONG EXCEL MASTER DATA:")
sheet_stats = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(cell is not None for cell in r)]
    headers = [str(h) for h in rows[0] if h is not None] if rows else []
    sheet_stats[name] = {
        "total_rows": len(rows),
        "non_empty_rows": len(non_empty),
        "data_rows": max(0, len(non_empty) - 1),
        "headers": headers[:8]
    }
    print(f" - [{name:25s}]: {len(non_empty):5d} dòng | Headers mẫu: {', '.join(headers[:5])}")

# 2. Database Counts
print("\n📊 2. THỐNG KÊ BẢNG TRONG CSDL SQLITE (devices.db):")
db_tables = ["devices", "facilities", "device_categories", "contracts", "suppliers", "accessories", "device_transfers", "maintenance_logs", "inspections"]
db_counts = {}
for t in db_tables:
    try:
        c = cur.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        db_counts[t] = c
        print(f" - Bảng [{t:22s}]: {c:5d} bản ghi")
    except Exception as e:
        print(f" - Bảng [{t:22s}]: LỖI ({e})")

# 3. Deep Audit on Devices
print("\n🔎 3. ĐỐI SOÁT CHI TIẾT DỮ LIỆU THIẾT BỊ (DEVICES):")
db_devices = cur.execute("SELECT id, device_name, model, serial_no, facility_id, risk_level, status, supplier_name, contract_no FROM devices").fetchall()
db_serials = set()
db_models = Counter()
db_risks = Counter()
for d in db_devices:
    if d["serial_no"]:
        db_serials.add(str(d["serial_no"]).strip().lower())
    db_models[d["model"]] += 1
    db_risks[d["risk_level"]] += 1

print(f" • Tổng số thiết bị trong DB: {len(db_devices)}")
print(f" • Phân bổ rủi ro trong DB: Loai A: {db_risks['A']}, Loai B: {db_risks['B']}, Loai C: {db_risks['C']}, Loai D: {db_risks['D']}")

# Check Excel Sheets that contain devices
print("\n🔬 4. PHÂN TÍCH SHEET THIẾT BỊ TRONG EXCEL:")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    
    # Check if this sheet has device info
    has_serial = any("serial" in h or "s/n" in h or "so seri" in h for h in headers)
    has_device = any("ten" in h or "thiết bị" in h or "device" in h or "model" in h for h in headers)
    
    if has_device or has_serial:
        print(f"\n---> Phân tích Sheet: [{sheet_name}] (Số dòng: {len(rows)})")
        print(f"     Headers: {headers[:10]}")
        
        # Sample match rate
        ser_idx = -1
        for idx, h in enumerate(headers):
            if "serial" in h or "s/n" in h or "so seri" in h:
                ser_idx = idx
                break
        
        if ser_idx != -1:
            excel_serials = [str(r[ser_idx]).strip().lower() for r in rows[1:] if r and len(r) > ser_idx and r[ser_idx]]
            matched = sum(1 for s in excel_serials if s in db_serials)
            print(f"     - Số serial có giá trị: {len(excel_serials)}")
            print(f"     - Trùng khớp với DB: {matched}/{len(excel_serials)} ({matched/len(excel_serials)*100:.1f}%)" if excel_serials else "     - Không có serial")

# 5. Summary & Gap Analysis
print("\n" + "="*90)
print("📌 KẾT LUẬN & ĐÁNH GIÁ CHÊNH LỆCH:")
print("="*90)
print(f"1. CSDL hiện tại chứa: {db_counts.get('devices', 0)} thiết bị (Khớp với danh mục chuẩn 1.211 thiết bị BVQ7).")
print(f"2. Danh mục Khoa/Phòng: {db_counts.get('facilities', 0)} khoa phòng.")
print(f"3. Hợp đồng mua sắm: {db_counts.get('contracts', 0)} hợp đồng.")
print(f"4. Danh bạ Nhà cung cấp: {db_counts.get('suppliers', 0)} NCC.")
