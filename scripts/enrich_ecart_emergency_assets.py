import sqlite3
import openpyxl
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent.parent))

db_path = Path(r"C:\Users\tantt\Downloads\medical-device-app\database\devices.db")
ecart_excel = Path(r"C:\Users\tantt\Downloads\TA5. VỊ TRÍ KHOA PHÒNG - XE ECART.xlsx")

print("🏥 TÍCH HỢP HỆ THỐNG XE CẤP CỨU DI ĐỘNG (E-CART) & VỊ TRÍ KHOA PHÒNG TA5:\n" + "=" * 75)

conn = sqlite3.connect(db_path)
cur = conn.cursor()

# 1. Tạo bảng emergency_carts
cur.execute("""
    CREATE TABLE IF NOT EXISTS emergency_carts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cart_code TEXT UNIQUE NOT NULL,
        location_floor TEXT NOT NULL,
        zone TEXT,
        room_no TEXT,
        phone_ext TEXT,
        department_name TEXT NOT NULL,
        defibrillator_sn TEXT,
        suction_unit_sn TEXT,
        status TEXT DEFAULT 'READY', -- READY, IN_USE, INSPECTION_DUE
        notes TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
""")

# 2. Đọc dữ liệu từ Sheet [VỊ TRÍ XE ECART]
wb = openpyxl.load_workbook(ecart_excel, data_only=True)
ws = wb['VỊ TRÍ XE ECART']

ecarts = []
for r in range(3, ws.max_row + 1):
    stt = ws.cell(r, 1).value
    floor = ws.cell(r, 2).value
    zone = ws.cell(r, 3).value
    room = ws.cell(r, 4).value
    phone = ws.cell(r, 5).value
    dept = ws.cell(r, 6).value
    qty = ws.cell(r, 7).value
    
    if dept or room:
        floor_str = str(floor or 'TẦNG TRỆT').strip()
        cart_code = f"ECART-Q7-0{len(ecarts)+1}"
        dept_str = str(dept or '').strip()
        room_str = str(room or '').strip().replace('.0', '')
        phone_str = str(phone or '').strip().replace('.0', '')
        zone_str = str(zone or '').strip()
        
        ecarts.append((
            cart_code, floor_str, zone_str, room_str, phone_str, dept_str,
            f"Xe cấp cứu di động tiêu chuẩn ACLS tại {dept_str}"
        ))

cur.execute("DELETE FROM emergency_carts")
for ec in ecarts:
    cur.execute("""
        INSERT INTO emergency_carts (cart_code, location_floor, zone, room_no, phone_ext, department_name, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, ec)

conn.commit()
print(f"✅ Đã nạp thành công {len(ecarts)} Xe Cấp Cứu Di Động (E-Cart) vào Cơ Sở Dữ Liệu:")
for idx, ec in enumerate(ecarts, 1):
    print(f"   {idx:02d}. [{ec[0]}] {ec[1]} - Khu {ec[2]} (Phòng {ec[3]} | Ext: {ec[4]}): {ec[5]}")

# 3. Cập nhật Semantica Graph Engine
from app.semantica_engine import semantica_engine, GraphNode, GraphEdge
semantica_engine._build_knowledge_graph()

for ec in ecarts:
    cart_node_id = f"ECART-{ec[0]}"
    semantica_engine.add_node(GraphNode(cart_node_id, "EmergencyCart", f"Xe Cấp Cứu {ec[5]}", {
        "cart_code": ec[0],
        "floor": ec[1],
        "room": ec[3],
        "phone": ec[4],
        "department": ec[5]
    }))

stats = semantica_engine.get_graph_stats()
print(f"\n🕸️ SEMANTICA AGI KNOWLEDGE GRAPH SAU KHI TÍCH HỢP E-CART:")
print(f"  • Tổng số Thực Thể (Nodes): {stats['total_nodes']:,} nodes")
print(f"  • Tổng số Mối Quan Hệ (Edges): {stats['total_edges']:,} edges")
print(f"  • Phân bổ Nodes: {stats['node_distribution']}")

conn.close()
